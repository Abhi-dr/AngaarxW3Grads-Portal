from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import IntegrityError
from django.db.models import Q

from student.event_models import Event, CertificateTemplate, Certificate
from accounts.models import CustomUser
from instructor.api.permissions import IsInstructorSubmission
from administration.api.serializers.event_serializers import (
    EventSerializer, CertificateTemplateSerializer,
    CertificateSerializer, StudentSearchSerializer,
)


class CertificateTemplateInstructorViewSet(viewsets.ModelViewSet):
    """
    Instructor access for certificate HTML templates.
    """
    queryset = CertificateTemplate.objects.all().order_by('-created_at')
    serializer_class = CertificateTemplateSerializer
    permission_classes = [IsInstructorSubmission]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Template deleted.'}, status=status.HTTP_200_OK)


class EventInstructorViewSet(viewsets.ModelViewSet):
    """
    Instructor access for Events.
    """
    queryset = Event.objects.all().select_related('certificate_template').order_by('-start_date')
    serializer_class = EventSerializer
    permission_classes = [IsInstructorSubmission]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Event deleted.'}, status=status.HTTP_200_OK)


class CertificateInstructorViewSet(viewsets.ModelViewSet):
    """
    Instructor access for Certificates.
    """
    serializer_class = CertificateSerializer
    permission_classes = [IsInstructorSubmission]

    def get_queryset(self):
        qs = Certificate.objects.select_related('event', 'student').order_by('-issued_date')
        event_id = self.request.query_params.get('event_id')
        if event_id:
            qs = qs.filter(event_id=event_id)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            try:
                cert = serializer.save()
                return Response({'success': True, 'data': self.get_serializer(cert).data}, status=status.HTTP_201_CREATED)
            except IntegrityError:
                return Response({'success': False, 'error': 'A certificate for this student and event already exists.'}, status=status.HTTP_409_CONFLICT)
        errors = serializer.errors
        if 'non_field_errors' in errors:
            return Response({'success': False, 'error': 'A certificate for this student and event already exists.'}, status=status.HTTP_409_CONFLICT)
        return Response({'success': False, 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Certificate deleted.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='mass-assign')
    def mass_assign(self, request):
        event_id = request.data.get('event_id')
        student_ids = request.data.get('student_ids', [])

        if not event_id:
            return Response({'success': False, 'error': 'event_id is required.'}, status=400)
        if not student_ids or not isinstance(student_ids, list):
            return Response({'success': False, 'error': 'student_ids must be a non-empty list.'}, status=400)

        try:
            event = Event.objects.get(pk=event_id)
        except Event.DoesNotExist:
            return Response({'success': False, 'error': 'Event not found.'}, status=404)

        students = CustomUser.objects.filter(pk__in=student_ids)
        created_count = 0
        skipped_count = 0
        errors = []

        for student in students:
            try:
                Certificate.objects.create(event=event, student=student)
                created_count += 1
            except IntegrityError:
                skipped_count += 1
            except Exception as e:
                errors.append(f"{student.email}: {str(e)}")

        return Response({
            'success': True,
            'created': created_count,
            'skipped_duplicates': skipped_count,
            'errors': errors,
            'message': f'{created_count} certificate(s) issued. {skipped_count} duplicate(s) skipped.'
        })

    @action(detail=False, methods=['get'], url_path='search-students')
    def search_students(self, request):
        q = request.query_params.get('q', '').strip()
        if not q or len(q) < 2:
            return Response({'success': False, 'error': 'Query must be at least 2 characters.'}, status=400)

        students = CustomUser.objects.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q),
        ).order_by('first_name')[:20]

        serializer = StudentSearchSerializer(students, many=True)
        return Response({'success': True, 'students': serializer.data})

    @action(detail=False, methods=['post'], url_path='mass-assign-csv')
    def mass_assign_csv(self, request):
        import csv
        import io

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'success': False, 'error': 'No file uploaded. Use field name "file".'}, status=400)

        filename = uploaded.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                text = uploaded.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
            elif filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                except ImportError:
                    return Response({'success': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=500)
                wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
            else:
                return Response({'success': False, 'error': 'Unsupported file type. Upload a .csv or .xlsx file.'}, status=400)
        except Exception as e:
            return Response({'success': False, 'error': f'Could not parse file: {str(e)}'}, status=400)

        if not rows:
            return Response({'success': False, 'error': 'File is empty or has no data rows.'}, status=400)

        first = rows[0]
        if 'email' not in first or 'event_id' not in first:
            return Response({'success': False, 'error': 'File must have columns: email, event_id'}, status=400)

        created_count = 0
        skipped_count = 0
        not_found = []
        errors = []

        event_cache = {}
        student_cache = {}

        for i, row in enumerate(rows, start=2):
            email = row.get('email', '').strip()
            event_id_raw = row.get('event_id', '').strip()

            if not email or not event_id_raw:
                continue
            try:
                event_id = int(float(event_id_raw))
            except ValueError:
                errors.append(f"Row {i}: invalid event_id")
                continue

            event = event_cache.get(event_id)
            if event is None:
                event = Event.objects.filter(pk=event_id).first()
                event_cache[event_id] = event
            if not event:
                not_found.append(email)
                continue

            student = student_cache.get(email)
            if student is None:
                student = CustomUser.objects.filter(email__iexact=email).first()
                student_cache[email] = student
            if not student:
                not_found.append(email)
                continue

            try:
                Certificate.objects.create(event=event, student=student)
                created_count += 1
            except IntegrityError:
                skipped_count += 1
            except Exception as e:
                errors.append(f"{email}: {str(e)}")

        return Response({
            'success': True,
            'created': created_count,
            'skipped_duplicates': skipped_count,
            'not_found': not_found,
            'errors': errors,
            'message': f'{created_count} certificate(s) issued. {skipped_count} duplicate(s) skipped.'
        })

    @action(detail=True, methods=['patch'], url_path='toggle-approved')
    def toggle_approved(self, request, pk=None):
        cert = self.get_object()
        cert.approved = not cert.approved
        cert.save(update_fields=['approved'])
        return Response({'success': True, 'message': f'Certificate {"approved" if cert.approved else "unapproved"}.'})

    @action(detail=False, methods=['post'], url_path='approve-all')
    def approve_all(self, request):
        event_id = request.data.get('event_id')
        if not event_id:
            return Response({'success': False, 'error': 'event_id is required.'}, status=400)

        pending_qs = Certificate.objects.filter(event_id=event_id, approved=False)
        total_count = Certificate.objects.filter(event_id=event_id).count()
        approved_now = pending_qs.count()
        pending_qs.update(approved=True)

        return Response({
            'success': True,
            'approved_now': approved_now,
            'total_count': total_count,
            'message': f'{approved_now} certificate(s) approved.'
        })
