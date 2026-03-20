from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db import IntegrityError
from django.db.models import Q

from student.event_models import Event, CertificateTemplate, Certificate
from accounts.models import CustomUser
from administration.api.permissions import IsAdministrator, IsAdministratorOrInstructor
from administration.api.serializers.event_serializers import (
    EventSerializer, CertificateTemplateSerializer,
    CertificateSerializer, StudentSearchSerializer
)


# ─────────────────────────────────────────────────────────────
# Certificate Template CRUD
# ─────────────────────────────────────────────────────────────

class CertificateTemplateAdminViewSet(viewsets.ModelViewSet):
    """
    CRUD for certificate HTML templates.
    GET/POST  /api/v1/certificate-templates/
    GET/PUT/PATCH/DELETE /api/v1/certificate-templates/<pk>/
    """
    queryset = CertificateTemplate.objects.all().order_by('-created_at')
    serializer_class = CertificateTemplateSerializer
    permission_classes = [IsAdministrator]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Template deleted.'}, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# Event CRUD
# ─────────────────────────────────────────────────────────────

class EventAdminViewSet(viewsets.ModelViewSet):
    """
    CRUD for Events.
    GET/POST  /api/v1/events/
    GET/PUT/PATCH/DELETE /api/v1/events/<pk>/
    """
    queryset = Event.objects.all().select_related('certificate_template').order_by('-start_date')
    serializer_class = EventSerializer
    permission_classes = [IsAdministrator]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Event deleted.'}, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# Certificate CRUD + Mass Assign + Student Search
# ─────────────────────────────────────────────────────────────

class CertificateAdminViewSet(viewsets.ModelViewSet):
    """
    CRUD for Certificates.
    GET  /api/v1/certificates/?event_id=<id>          - list for an event
    POST /api/v1/certificates/                        - create single certificate
    PATCH /api/v1/certificates/<pk>/                  - partial update
    DELETE /api/v1/certificates/<pk>/                 - delete

    POST /api/v1/certificates/mass-assign/
         Body: { event_id, student_ids: [int, ...] }
         Bulk-creates certificates for the given students.

    GET /api/v1/certificates/search-students/
        ?q=<name|email>   returns matching students
    """
    serializer_class = CertificateSerializer
    permission_classes = [IsAdministrator]

    def get_queryset(self):
        qs = Certificate.objects.select_related('event', 'student').order_by('-issued_date')
        event_id = self.request.query_params.get('event_id')
        if event_id:
            qs = qs.filter(event_id=event_id)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            try:
                cert = serializer.save()
                return Response({'success': True, 'data': self.get_serializer(cert).data}, status=status.HTTP_201_CREATED)
            except IntegrityError:
                return Response({
                    'success': False,
                    'error': 'A certificate for this student and event already exists.'
                }, status=status.HTTP_409_CONFLICT)
        # Catch DRF's unique_together validation error and return a friendly message
        errors = serializer.errors
        if 'non_field_errors' in errors:
            return Response({
                'success': False,
                'error': 'A certificate for this student and event already exists.'
            }, status=status.HTTP_409_CONFLICT)
        return Response({'success': False, 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'success': True, 'message': 'Certificate deleted.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='mass-assign')
    def mass_assign(self, request):
        """
        POST /api/v1/certificates/mass-assign/
        { event_id: int, student_ids: [int, ...] }
        Bulk-creates certificates (skips duplicates).
        """
        event_id = request.data.get('event_id')
        student_ids = request.data.get('student_ids', [])

        if not event_id:
            return Response({'success': False, 'error': 'event_id is required.'}, status=400)
        if not student_ids or not isinstance(student_ids, list):
            return Response({'success': False, 'error': 'student_ids must be a non-empty list.'}, status=400)

        try:
            event = Event.objects.get(pk=event_id)
        except Event.DoesNotExist:
            return Response({'success': False, 'error': 'Event not found.'}, status=404)

        students = CustomUser.objects.filter(pk__in=student_ids)
        created_count = 0
        skipped_count = 0
        errors = []

        for student in students:
            try:
                Certificate.objects.create(event=event, student=student)
                created_count += 1
            except IntegrityError:
                skipped_count += 1
            except Exception as e:
                errors.append(f"{student.email}: {str(e)}")

        return Response({
            'success': True,
            'created': created_count,
            'skipped_duplicates': skipped_count,
            'errors': errors,
            'message': f'{created_count} certificate(s) issued. {skipped_count} duplicate(s) skipped.'
        })

    @action(detail=False, methods=['get'], url_path='search-students')
    def search_students(self, request):
        """
        GET /api/v1/certificates/search-students/?q=<name|email>
        Returns matching students (role=student).
        """
        q = request.query_params.get('q', '').strip()
        if not q or len(q) < 2:
            return Response({'success': False, 'error': 'Query must be at least 2 characters.'}, status=400)

        students = CustomUser.objects.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q),
        ).order_by('first_name')[:20]

        serializer = StudentSearchSerializer(students, many=True)
        return Response({'success': True, 'students': serializer.data})

    @action(detail=False, methods=['post'], url_path='mass-assign-csv')
    def mass_assign_csv(self, request):
        """
        POST /api/v1/certificates/mass-assign-csv/
        Multipart file upload: field name = 'file'
        Accepted formats: .csv or .xlsx
        Required columns: email, event_id
        Optional column: issued_date (YYYY-MM-DD)
        Assigns certificates based on each row's event_id + email.
        """
        import csv
        import io

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'success': False, 'error': 'No file uploaded. Use field name "file".'}, status=400)

        filename = uploaded.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                text = uploaded.read().decode('utf-8-sig')  # handles BOM
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
            elif filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                except ImportError:
                    return Response({'success': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=500)
                wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
            else:
                return Response({'success': False, 'error': 'Unsupported file type. Upload a .csv or .xlsx file.'}, status=400)
        except Exception as e:
            return Response({'success': False, 'error': f'Could not parse file: {str(e)}'}, status=400)

        if not rows:
            return Response({'success': False, 'error': 'File is empty or has no data rows.'}, status=400)

        # Validate required columns
        first = rows[0]
        if 'email' not in first or 'event_id' not in first:
            return Response({'success': False, 'error': 'File must have columns: email, event_id'}, status=400)

        created_count = 0
        skipped_count = 0
        not_found = []
        errors = []

        # Cache events and students to avoid N+1 queries
        event_cache = {}
        student_cache = {}

        for i, row in enumerate(rows, start=2):
            email = row.get('email', '').strip()
            event_id_raw = row.get('event_id', '').strip()

            if not email or not event_id_raw:
                continue
            try:
                event_id = int(float(event_id_raw))
            except (ValueError, TypeError):
                errors.append(f'Row {i}: invalid event_id "{event_id_raw}"')
                continue

            # Get or cache event
            if event_id not in event_cache:
                try:
                    event_cache[event_id] = Event.objects.get(pk=event_id)
                except Event.DoesNotExist:
                    event_cache[event_id] = None
            event = event_cache[event_id]
            if not event:
                errors.append(f'Row {i}: event_id {event_id} not found')
                continue

            # Get or cache student
            if email not in student_cache:
                try:
                    student_cache[email] = CustomUser.objects.get(email=email)
                except CustomUser.DoesNotExist:
                    student_cache[email] = None
            student = student_cache[email]
            if not student:
                not_found.append(email)
                continue

            try:
                Certificate.objects.create(event=event, student=student)
                created_count += 1
            except IntegrityError:
                skipped_count += 1

        return Response({
            'success': True,
            'created': created_count,
            'skipped_duplicates': skipped_count,
            'not_found_emails': not_found,
            'errors': errors,
            'message': f'{created_count} issued, {skipped_count} duplicates skipped, {len(not_found)} emails not found.'
        })

    @action(detail=False, methods=['post'], url_path='approve-all')
    def approve_all(self, request):
        """
        POST /api/v1/certificates/approve-all/
        { event_id: int }
        Marks all pending certificates for the given event as approved.
        """
        event_id = request.data.get('event_id')
        if not event_id:
            return Response({'success': False, 'error': 'event_id is required.'}, status=400)

        try:
            event = Event.objects.get(pk=event_id)
        except Event.DoesNotExist:
            return Response({'success': False, 'error': 'Event not found.'}, status=404)

        pending_qs = Certificate.objects.filter(event=event, approved=False)
        approved_now = pending_qs.update(approved=True)
        total_count = Certificate.objects.filter(event=event).count()

        return Response({
            'success': True,
            'approved_now': approved_now,
            'total': total_count,
            'message': f'{approved_now} certificate(s) approved.'
        })

    @action(detail=True, methods=['patch'], url_path='toggle-approved')
    def toggle_approved(self, request, pk=None):
        """PATCH /api/v1/certificates/<pk>/toggle-approved/ — flip approved flag."""
        cert = self.get_object()
        cert.approved = not cert.approved
        cert.save(update_fields=['approved'])
        return Response({
            'success': True,
            'approved': cert.approved,
            'message': f'Certificate {"approved" if cert.approved else "unapproved"}.'
        })
